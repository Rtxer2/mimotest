import io
from datetime import datetime
from sqlalchemy.orm import Session

from app.models.order import Order, OrderItem
from app.models.production import ProductionOrder
from app.models.quality import QualityInspection, QualityIssue
from app.models.inventory import Material, FinishedProduct
from app.models.customer import Customer


REPORT_TITLE = "SmartFactory ERP"
REPORT_DATE_FMT = "%Y-%m-%d %H:%M"


def _now_str():
    return datetime.now().strftime(REPORT_DATE_FMT)


def _style_xlsx_header(ws, col_count):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _style_xlsx_data(ws, row_count, col_count):
    from openpyxl.styles import Alignment, Border, Side, PatternFill
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    for row in range(2, row_count + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if row % 2 == 0:
                cell.fill = alt_fill


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val.encode("utf-8")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 40)


def _pdf_styles():
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="ReportTitle",
        parent=styles["Title"],
        fontSize=18,
        spaceAfter=6,
        textColor=colors.HexColor("#2F5496"),
    ))
    styles.add(ParagraphStyle(
        name="ReportSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.grey,
        alignment=TA_CENTER,
        spaceAfter=20,
    ))
    styles.add(ParagraphStyle(
        name="SectionTitle",
        parent=styles["Heading2"],
        fontSize=13,
        textColor=colors.HexColor("#2F5496"),
        spaceBefore=16,
        spaceAfter=8,
    ))
    return styles


def _pdf_table_style():
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F7FB")]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ])


def _pdf_info_table_style():
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle
    return TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8EDF5")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ])


def _add_pdf_header(elements, title, styles):
    from reportlab.platypus import Spacer
    from reportlab.platypus import Paragraph
    elements.append(Paragraph(f"{REPORT_TITLE}", styles["ReportTitle"]))
    elements.append(Paragraph(f"{title}  |  {_now_str()}", styles["ReportSubtitle"]))


class ReportService:
    def __init__(self, db: Session):
        self.db = db

    def export_orders_xlsx(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        ws.append(["Order No", "Customer", "Status", "Total Amount", "Delivery Date", "Created At"])
        orders = self.db.query(Order).order_by(Order.created_at.desc()).all()
        for o in orders:
            ws.append([o.order_no, o.customer_id, o.status, float(o.total_amount or 0), str(o.delivery_date or "")[:10], str(o.created_at or "")[:19]])
        _style_xlsx_header(ws, 6)
        _style_xlsx_data(ws, len(orders) + 1, 6)
        _auto_width(ws)
        ws.freeze_panes = "A2"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def export_orders_pdf(self):
        from reportlab.platypus import SimpleDocTemplate, Table, Spacer
        from reportlab.lib.pagesizes import A4, landscape
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
        styles = _pdf_styles()
        elements = []
        _add_pdf_header(elements, "Orders Report", styles)
        orders = self.db.query(Order).order_by(Order.created_at.desc()).all()
        data = [["Order No", "Customer", "Status", "Total Amount", "Delivery Date", "Created At"]]
        for o in orders:
            data.append([o.order_no, str(o.customer_id), o.status, f"¥{float(o.total_amount or 0):,.2f}", str(o.delivery_date or "")[:10], str(o.created_at or "")[:10]])
        t = Table(data, repeatRows=1)
        t.setStyle(_pdf_table_style())
        elements.append(t)
        elements.append(Spacer(1, 12))
        from reportlab.platypus import Paragraph
        elements.append(Paragraph(f"Total: {len(orders)} orders", styles["Normal"]))
        doc.build(elements)
        buf.seek(0)
        return buf

    def export_production_xlsx(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Production"
        ws.append(["Order ID", "Related Order", "Status", "Workshop", "Planned Start", "Planned End"])
        orders = self.db.query(ProductionOrder).all()
        for o in orders:
            ws.append([o.id, o.order_id, o.status, o.assigned_workshop or "", str(o.planned_start or "")[:10], str(o.planned_end or "")[:10]])
        _style_xlsx_header(ws, 6)
        _style_xlsx_data(ws, len(orders) + 1, 6)
        _auto_width(ws)
        ws.freeze_panes = "A2"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def export_production_pdf(self):
        from reportlab.platypus import SimpleDocTemplate, Table, Spacer, Paragraph
        from reportlab.lib.pagesizes import A4, landscape
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
        styles = _pdf_styles()
        elements = []
        _add_pdf_header(elements, "Production Report", styles)
        orders = self.db.query(ProductionOrder).all()
        data = [["Order ID", "Related Order", "Status", "Workshop", "Planned Start", "Planned End"]]
        for o in orders:
            data.append([str(o.id), str(o.order_id), o.status, o.assigned_workshop or "", str(o.planned_start or "")[:10], str(o.planned_end or "")[:10]])
        t = Table(data, repeatRows=1)
        t.setStyle(_pdf_table_style())
        elements.append(t)
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Total: {len(orders)} production orders", styles["Normal"]))
        doc.build(elements)
        buf.seek(0)
        return buf

    def export_quality_xlsx(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Quality"
        ws.append(["ID", "Type", "Result", "Inspector", "Time", "Issues"])
        inspections = self.db.query(QualityInspection).all()
        for i in inspections:
            issue_count = self.db.query(QualityIssue).filter(QualityIssue.inspection_id == i.id).count()
            ws.append([i.id, i.inspection_type, i.result, i.inspector or "", str(i.inspect_time or "")[:19], issue_count])
        _style_xlsx_header(ws, 6)
        _style_xlsx_data(ws, len(inspections) + 1, 6)
        _auto_width(ws)
        ws.freeze_panes = "A2"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def export_quality_pdf(self):
        from reportlab.platypus import SimpleDocTemplate, Table, Spacer, Paragraph
        from reportlab.lib.pagesizes import A4, landscape
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
        styles = _pdf_styles()
        elements = []
        _add_pdf_header(elements, "Quality Report", styles)
        inspections = self.db.query(QualityInspection).all()
        data = [["ID", "Type", "Result", "Inspector", "Time", "Issues"]]
        for i in inspections:
            issue_count = self.db.query(QualityIssue).filter(QualityIssue.inspection_id == i.id).count()
            data.append([str(i.id), i.inspection_type, i.result, i.inspector or "", str(i.inspect_time or "")[:10], str(issue_count)])
        t = Table(data, repeatRows=1)
        t.setStyle(_pdf_table_style())
        elements.append(t)
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Total: {len(inspections)} inspections", styles["Normal"]))
        doc.build(elements)
        buf.seek(0)
        return buf

    def export_inventory_xlsx(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Materials"
        ws.append(["Code", "Name", "Unit", "Current Stock", "Safety Stock", "Status"])
        materials = self.db.query(Material).all()
        for m in materials:
            cur = float(m.current_stock or 0)
            safe = float(m.safety_stock or 0)
            status = "LOW" if safe > 0 and cur < safe else "OK"
            ws.append([m.code, m.name, m.unit, cur, safe, status])
        _style_xlsx_header(ws, 6)
        _style_xlsx_data(ws, len(materials) + 1, 6)
        from openpyxl.styles import Font, PatternFill
        for row in range(2, len(materials) + 2):
            if ws.cell(row=row, column=6).value == "LOW":
                ws.cell(row=row, column=6).font = Font(bold=True, color="FF0000")
                ws.cell(row=row, column=4).font = Font(color="FF0000")
        _auto_width(ws)
        ws.freeze_panes = "A2"

        ws2 = wb.create_sheet("Products")
        ws2.append(["SKU", "Name", "Current Stock", "Safety Stock", "Category", "Status"])
        products = self.db.query(FinishedProduct).all()
        for p in products:
            status = "LOW" if p.safety_stock > 0 and p.current_stock < p.safety_stock else "OK"
            ws2.append([p.sku, p.product_name, p.current_stock, p.safety_stock, p.category or "", status])
        _style_xlsx_header(ws2, 6)
        _style_xlsx_data(ws2, len(products) + 1, 6)
        for row in range(2, len(products) + 2):
            if ws2.cell(row=row, column=6).value == "LOW":
                ws2.cell(row=row, column=6).font = Font(bold=True, color="FF0000")
                ws2.cell(row=row, column=3).font = Font(color="FF0000")
        _auto_width(ws2)
        ws2.freeze_panes = "A2"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def export_inventory_pdf(self):
        from reportlab.platypus import SimpleDocTemplate, Table, Spacer, Paragraph
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
        styles = _pdf_styles()
        elements = []
        _add_pdf_header(elements, "Inventory Report", styles)

        elements.append(Paragraph("Materials", styles["SectionTitle"]))
        materials = self.db.query(Material).all()
        data = [["Code", "Name", "Unit", "Current Stock", "Safety Stock", "Status"]]
        for m in materials:
            cur = float(m.current_stock or 0)
            safe = float(m.safety_stock or 0)
            status = "LOW" if safe > 0 and cur < safe else "OK"
            data.append([m.code, m.name, m.unit, f"{cur:,.2f}", f"{safe:,.2f}", status])
        t = Table(data, repeatRows=1)
        ts = _pdf_table_style()
        t.setStyle(ts)
        elements.append(t)

        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Finished Products", styles["SectionTitle"]))
        products = self.db.query(FinishedProduct).all()
        pdata = [["SKU", "Name", "Current Stock", "Safety Stock", "Category", "Status"]]
        for p in products:
            status = "LOW" if p.safety_stock > 0 and p.current_stock < p.safety_stock else "OK"
            pdata.append([p.sku, p.product_name, str(p.current_stock), str(p.safety_stock), p.category or "", status])
        t2 = Table(pdata, repeatRows=1)
        t2.setStyle(_pdf_table_style())
        elements.append(t2)
        doc.build(elements)
        buf.seek(0)
        return buf

    def export_single_order_xlsx(self, order_id: int):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        order = self.db.query(Order).filter(Order.id == order_id).first()
        if not order:
            raise ValueError("Order not found")
        items = self.db.query(OrderItem).filter(OrderItem.order_id == order_id).all()
        customer = self.db.query(Customer).filter(Customer.id == order.customer_id).first()

        wb = Workbook()
        ws = wb.active
        ws.title = "Order Detail"

        title_font = Font(bold=True, size=16, color="2F5496")
        label_font = Font(bold=True, size=11)
        value_font = Font(size=11)
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        ws.merge_cells("A1:D1")
        ws["A1"] = f"Order: {order.order_no}"
        ws["A1"].font = title_font

        info = [
            ("Order No", order.order_no),
            ("Customer", customer.name if customer else str(order.customer_id)),
            ("Status", order.status),
            ("Total Amount", f"¥{float(order.total_amount or 0):,.2f}"),
            ("Delivery Date", str(order.delivery_date or "")[:10]),
            ("Remarks", order.remarks or ""),
            ("Created At", str(order.created_at or "")[:19]),
        ]
        for i, (label, value) in enumerate(info, start=3):
            ws.cell(row=i, column=1, value=label).font = label_font
            ws.cell(row=i, column=1).fill = PatternFill(start_color="E8EDF5", end_color="E8EDF5", fill_type="solid")
            ws.cell(row=i, column=1).border = thin_border
            ws.cell(row=i, column=2, value=str(value)).font = value_font
            ws.cell(row=i, column=2).border = thin_border

        item_start = len(info) + 4
        ws.cell(row=item_start, column=1, value="Order Items").font = Font(bold=True, size=13, color="2F5496")

        headers = ["Product", "Quantity", "Unit Price", "Amount", "Specs"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=item_start + 1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for i, item in enumerate(items, start=item_start + 2):
            amount = item.quantity * float(item.unit_price or 0)
            row_data = [item.product_name, item.quantity, float(item.unit_price or 0), amount, item.specs or ""]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin_border
                if col in (3, 4):
                    cell.number_format = '¥#,##0.00'

        total_row = item_start + len(items) + 2
        ws.cell(row=total_row, column=1, value="Total").font = label_font
        ws.cell(row=total_row, column=4, value=float(order.total_amount or 0)).font = label_font
        ws.cell(row=total_row, column=4).number_format = '¥#,##0.00'

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 25
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def export_single_order_pdf(self, order_id: int):
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors

        order = self.db.query(Order).filter(Order.id == order_id).first()
        if not order:
            raise ValueError("Order not found")
        items = self.db.query(OrderItem).filter(OrderItem.order_id == order_id).all()
        customer = self.db.query(Customer).filter(Customer.id == order.customer_id).first()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
        styles = _pdf_styles()
        elements = []

        _add_pdf_header(elements, f"Order {order.order_no}", styles)

        info = [
            ["Order No", order.order_no],
            ["Customer", customer.name if customer else str(order.customer_id)],
            ["Status", order.status],
            ["Total Amount", f"¥{float(order.total_amount or 0):,.2f}"],
            ["Delivery Date", str(order.delivery_date or "")[:10]],
            ["Created At", str(order.created_at or "")[:19]],
            ["Remarks", order.remarks or ""],
        ]
        t = Table(info, colWidths=[120, 350])
        t.setStyle(_pdf_info_table_style())
        elements.append(t)

        elements.append(Spacer(1, 24))
        elements.append(Paragraph("Order Items", styles["SectionTitle"]))

        item_data = [["Product", "Qty", "Unit Price", "Amount", "Specs"]]
        for item in items:
            amount = item.quantity * float(item.unit_price or 0)
            item_data.append([
                item.product_name,
                str(item.quantity),
                f"¥{float(item.unit_price or 0):,.2f}",
                f"¥{amount:,.2f}",
                item.specs or "",
            ])
        item_data.append(["Total", "", "", f"¥{float(order.total_amount or 0):,.2f}", ""])

        t2 = Table(item_data, colWidths=[130, 50, 90, 90, 120], repeatRows=1)
        ts = _pdf_table_style()
        ts.add("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8EDF5"))
        ts.add("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")
        t2.setStyle(ts)
        elements.append(t2)

        doc.build(elements)
        buf.seek(0)
        return buf

    def export(self, report_type: str, format: str):
        method_name = f"export_{report_type}_{format}"
        method = getattr(self, method_name, None)
        if not method:
            raise ValueError(f"Unsupported: {report_type}.{format}")
        return method()
